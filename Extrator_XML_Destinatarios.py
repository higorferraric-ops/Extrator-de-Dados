import os
import xml.etree.ElementTree as ET
import pandas as pd
import customtkinter as ctk
import tkinter as tk
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading # Import threading
from tkinter import filedialog, messagebox
import openpyxl

# --- Configuração Visual (Mantendo o padrão do seu ecossistema) ---
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

COLORS = {
    "bg": "#0f172a",       # Slate 900
    "surface": "#1e293b",  # Slate 800
    "surface2": "#334155", # Slate 700
    "border": "#334155",
    "accent": "#3b82f6",   # Blue 500
    "green": "#10b981",    # Emerald 500
    "yellow": "#f59e0b",
    "text": "#f8fafc",     # Slate 50
    "text_dark": "#1e293b",
    "muted": "#94a3b8",    # Slate 400
}

def extrair_dados_recursivo(elemento, prefixo=""):
    """
    Função recursiva para extrair todas as tags dentro de <dest>,
    mesmo as aninhadas, criando colunas achatadas.
    """
    dados = {}
    for filho in elemento:
        tag = filho.tag.split('}')[-1]  # Remove namespace
        nome_coluna = f"{prefixo}{tag}" if prefixo else tag
        
        if len(filho) == 0:  # É uma tag folha (tem valor direto)
            dados[nome_coluna] = filho.text if filho.text is not None else "" # Garante string vazia, não None
        else:  # Tem sub-tags (ex: enderDest)
            dados.update(extrair_dados_recursivo(filho, f"{nome_coluna}_"))
    return dados

def processar_xmls(caminhos, extrair_dest=True, extrair_emit=True, extrair_transp=True, extrair_infcpl=True, progress_callback=None):
    todos_dados = []
    total_arquivos = len(caminhos)
    
    for i, caminho in enumerate(caminhos):
        if progress_callback:
            progress_callback(i + 1, total_arquivos, os.path.basename(caminho))
        try:
            tree = ET.parse(caminho)
            root = tree.getroot()
            
            # Identifica namespaces
            ns = ""
            if '}' in root.tag:
                ns = root.tag.split('}')[0] + '}'
            
            registro = {}
            registro['Arquivo_Origem'] = os.path.basename(caminho)

            # 1. Extração do Destinatário
            if extrair_dest:
                dest = root.find(f'.//{ns}dest')
                if dest is not None:
                    dados_dest = extrair_dados_recursivo(dest, prefixo="Dest_")
                    registro.update(dados_dest)

            # 2. Extração do Emitente
            if extrair_emit:
                emit = root.find(f'.//{ns}emit')
                if emit is not None:
                    dados_emit = extrair_dados_recursivo(emit, prefixo="Emit_")
                    registro.update(dados_emit)

            # 3. Extração da Transportadora
            if extrair_transp:
                transp = root.find(f'.//{ns}transp')
                if transp is not None:
                    dados_transp = extrair_dados_recursivo(transp, prefixo="Transp_")
                    registro.update(dados_transp)

            # 4. Informações Complementares
            if extrair_infcpl:
                infAdic_elem = root.find(f'.//{ns}infAdic')
                if infAdic_elem is not None:
                    infCpl_elem = infAdic_elem.find(f'{ns}infCpl')
                    registro['Informacoes_Complementares'] = infCpl_elem.text if infCpl_elem is not None else ""
                else:
                    registro['Informacoes_Complementares'] = ""
            
            if len(registro) > 1: # Se extraiu algo além do nome do arquivo
                todos_dados.append(registro)
                    
        except Exception as e:
            print(f"Erro ao processar {caminho}: {e}")
            
    return todos_dados

class XMLToExcelApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Extração de Dados - XML")
        self.geometry("1000x700")
        self.configure(fg_color=COLORS["bg"])
        
        self.arquivos_selecionados = []
        self.arquivos_para_processar = []
        self.dados_extraidos = []

        # Configuração de Fontes Modernas
        self.font_main = ("Segoe UI", 13)
        self.font_bold = ("Segoe UI", 13, "bold")
        self.font_title = ("Segoe UI", 18, "bold")
        self.font_tag = ("Segoe UI", 10, "bold")

        self._setup_ui()

    def _setup_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=70, corner_radius=0)
        header.pack(fill="x", side="top")
        
        ctk.CTkLabel(header, text="EXTRATOR XML", 
                     font=self.font_tag,
                     fg_color=COLORS["accent"], text_color="white",
                     corner_radius=6).pack(side="left", padx=(30, 10))
        
        ctk.CTkLabel(header, text="Extração de Dados - XML",
                     font=self.font_title,
                     text_color=COLORS["text"]).pack(side="left")

        # Main Content Area
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=60, pady=40)

        # Drop Area (Simulada)
        self.drop_card = ctk.CTkFrame(self.main_frame, fg_color=COLORS["surface"], 
                                      border_color=COLORS["border"], border_width=1, corner_radius=15)
        self.drop_card.pack(fill="both", expand=True)
        
        # Painel de Opções
        self.options_frame = ctk.CTkFrame(self.drop_card, fg_color="transparent")
        self.options_frame.pack(pady=(30, 0))

        ctk.CTkLabel(self.options_frame, text="TAGS PARA EXTRAÇÃO:", 
                     font=self.font_tag, text_color=COLORS["accent"]).grid(row=0, column=0, columnspan=2, pady=(0, 15))

        self.check_dest = ctk.CTkCheckBox(self.options_frame, text="Clientes", font=self.font_main, corner_radius=6)
        self.check_dest.grid(row=1, column=0, padx=20, pady=5, sticky="w")

        self.check_emit = ctk.CTkCheckBox(self.options_frame, text="Fornecedores", font=self.font_main, corner_radius=6)
        self.check_emit.grid(row=1, column=1, padx=20, pady=5, sticky="w")

        self.check_transp = ctk.CTkCheckBox(self.options_frame, text="Transportadoras", font=self.font_main, corner_radius=6)
        self.check_transp.grid(row=2, column=0, padx=20, pady=5, sticky="w")

        self.check_infcpl = ctk.CTkCheckBox(self.options_frame, text="Info. Complementares", font=self.font_main, corner_radius=6)
        self.check_infcpl.grid(row=2, column=1, padx=20, pady=5, sticky="w")

        # Separador visual
        ctk.CTkFrame(self.drop_card, fg_color=COLORS["border"], height=1).pack(fill="x", padx=120, pady=30)

        self.info_label = ctk.CTkLabel(self.drop_card, text="Selecione os arquivos XML para começar",
                                       font=self.font_main, text_color=COLORS["muted"])
        self.info_label.pack(pady=(40, 20))

        btn_frame = ctk.CTkFrame(self.drop_card, fg_color="transparent")
        btn_frame.pack(pady=(10, 40))

        self.select_files_btn = ctk.CTkButton(btn_frame, text="Selecionar Arquivos", 
                                              command=self._selecionar_arquivos,
                                              fg_color=COLORS["accent"], hover_color="#2563eb",
                                              height=42, corner_radius=10, font=self.font_bold)
        self.select_files_btn.pack(side="left", padx=10)
        
        self.select_folder_btn = ctk.CTkButton(btn_frame, text="Selecionar Pasta", 
                                               command=self._selecionar_pasta,
                                               fg_color="transparent", border_width=1, border_color=COLORS["accent"],
                                               height=42, corner_radius=10, font=self.font_bold)
        self.select_folder_btn.pack(side="left", padx=10)

        # Status Bar
        self.status_label = ctk.CTkLabel(self, text="Aguardando arquivos...", 
                                         font=ctk.CTkFont(size=11), text_color=COLORS["muted"])
        self.status_label.pack(side="bottom", pady=10)

        # Action Button (Inicia oculto)
        self.process_btn = ctk.CTkButton(self.main_frame, text="Extrair Informações via Excel", 
                                         command=self._salvar_excel,
                                         fg_color=COLORS["green"], hover_color="#059669",
                                         font=self.font_bold, height=52, corner_radius=12)

        # Botão para Limpar (Inicia oculto)
        self.reset_btn = ctk.CTkButton(self.main_frame, text="Limpar / Novo Processamento", 
                                         command=self._limpar_dados,
                                         fg_color="transparent", border_width=1, 
                                         border_color=COLORS["yellow"], text_color=COLORS["yellow"],
                                         font=self.font_bold, height=40, corner_radius=12)

    def _selecionar_arquivos(self):
        paths = filedialog.askopenfilenames(filetypes=[("Arquivos XML", "*.xml")])
        if paths:
            self._carregar_e_processar(list(paths))

    def _selecionar_pasta(self):
        folder = filedialog.askdirectory()
        if folder:
            paths = []
            # Busca recursiva em todas as subpastas para identificar todos os arquivos XML
            for root_dir, _, files in os.walk(folder):
                for f in files:
                    if f.lower().endswith('.xml'):
                        paths.append(os.path.join(root_dir, f))
            
            if not paths:
                messagebox.showwarning("Aviso", "Nenhum arquivo XML encontrado nesta pasta.")
                return
            self._carregar_e_processar(paths)

    def _carregar_e_processar(self, caminhos):
        """Apenas armazena os caminhos e prepara a UI para o processamento"""
        self.arquivos_para_processar = caminhos
        self.process_btn.pack_forget()
        self.reset_btn.pack_forget()
        
        self.info_label.configure(text=f"{len(caminhos)} arquivo(s) carregados", text_color=COLORS["text"])
        self.process_btn.pack(pady=(20, 0), fill="x")
        self.reset_btn.pack(pady=(10, 0), fill="x")
        self.status_label.configure(text="Ajuste as tags acima e clique em 'Gerar Excel' para extrair os dados.")

    def _limpar_dados(self):
        # Resetar variáveis de dados
        self.dados_extraidos = []
        self.arquivos_para_processar = []
        
        # Resetar interface
        self.info_label.configure(text="Selecione os arquivos XML para começar", text_color=COLORS["muted"])
        self.status_label.configure(text="Aguardando arquivos...", text_color=COLORS["muted"])
        self.process_btn.configure(state="normal", text="Extrair Informações via Excel")
        self.process_btn.pack_forget()
        self.reset_btn.pack_forget()

    def _salvar_excel(self):
        if not self.arquivos_para_processar:
            return

        # UI Feedback - Desabilita botões e avisa sobre o processamento
        self.process_btn.configure(state="disabled", text="Extraindo dados...")
        self.info_label.configure(text="Processando XMLs...", text_color=COLORS["yellow"])

        # Captura as configurações no exato momento do clique (Dinâmico)
        extrair_dest_f = bool(self.check_dest.get())
        extrair_emit_f = bool(self.check_emit.get())
        extrair_transp_f = bool(self.check_transp.get())
        extrair_infcpl_f = bool(self.check_infcpl.get())

        def run_extraction():
            def update_status(atual, total, arquivo):
                self.after(0, lambda: self.status_label.configure(text=f"Processando ({atual}/{total}): {arquivo}"))

            # Realiza a extração real dos arquivos
            self.dados_extraidos = processar_xmls(
                self.arquivos_para_processar, 
                extrair_dest=extrair_dest_f, 
                extrair_emit=extrair_emit_f,
                extrair_transp=extrair_transp_f,
                extrair_infcpl=extrair_infcpl_f,
                progress_callback=update_status
            )
            self.after(0, self._concluir_e_abrir_dialogo)

        threading.Thread(target=run_extraction, daemon=True).start()

    def _concluir_e_abrir_dialogo(self):
        self.process_btn.configure(state="normal", text="Extrair Informações via Excel")
        self.info_label.configure(text=f"{len(self.arquivos_para_processar)} arquivo(s) carregados", text_color=COLORS["text"])
        
        if not self.dados_extraidos:
            messagebox.showerror("Erro", "Nenhum dado encontrado com as tags selecionadas.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Extração de dados.xlsx"
        )

        if not save_path:
            return

        try:
            df = pd.DataFrame(self.dados_extraidos)
            
            # Reorganizar colunas: Arquivo_Origem por último
            cols = [c for c in df.columns if c != 'Arquivo_Origem'] + ['Arquivo_Origem']
            df = df[cols]

            # --- Agrupamento por CNPJ ou CPF (Deduplicação) ---
            # Identificadores comuns em NFe para evitar linhas repetidas do mesmo cliente/fornecedor
            id_cols = ['Dest_CNPJ', 'Dest_CPF']
            for col in id_cols:
                if col in df.columns:
                    # Filtramos apenas as linhas onde o identificador está preenchido
                    # para evitar remover registros sem identificação (ex: vendas para consumidor final sem CPF)
                    mask_filled = (df[col].notna()) & (df[col].astype(str).str.strip() != "")
                    df_filled = df[mask_filled].drop_duplicates(subset=[col], keep='first')
                    df_empty = df[~mask_filled]
                    df = pd.concat([df_filled, df_empty], ignore_index=True)

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Extração de dados')
                
                worksheet = writer.sheets['Extração de dados']
                
                # Estilos
                fill_header = PatternFill("solid", fgColor="1E3A5F")
                fill_zebra = PatternFill("solid", fgColor="F1F5F9")
                font_header = Font(color="FFFFFF", bold=True)
                thin_border = Border(
                    left=Side(style='thin', color="CCCCCC"),
                    right=Side(style='thin', color="CCCCCC"),
                    top=Side(style='thin', color="CCCCCC"),
                    bottom=Side(style='thin', color="CCCCCC")
                )

                # Aplicar estilo ao Cabeçalho
                for cell in worksheet[1]:
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # Aplicar estilos ao corpo e Auto-ajuste
                for idx, col in enumerate(df.columns):
                    max_val_len = df[col].astype(str).str.len().max()
                    if pd.isna(max_val_len): max_val_len = 0
                    max_len = max(int(max_val_len), len(str(col))) + 2
                    
                    col_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[col_letter].width = min(max_len, 50)

                    # Zebra e Bordas nas linhas
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=idx + 1)
                        cell.border = thin_border
                        if row_idx % 2 == 0:
                            cell.fill = fill_zebra

                # Congelar primeira linha
                worksheet.freeze_panes = "A2"

            messagebox.showinfo("Sucesso", f"Excel gerado com sucesso!\nSalvo em: {save_path}")
            self.status_label.configure(text=f"Extração finalizada: {len(df)} registros únicos exportados.")

        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Ocorreu um erro ao gerar o Excel: {e}")

if __name__ == "__main__":
    app = XMLToExcelApp()
    app.mainloop()